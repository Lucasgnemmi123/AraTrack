from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, flash, session
from datetime import datetime
import os
import sys
from db_manager import DBManager
from maestras_manager import MaestrasManager
from pdf_generator import PDFGenerator
from auth_manager import AuthManager, login_required
import rendiciones_manager

app = Flask(__name__)
app.secret_key = 'aratrack-pro-2025-secure-key'
db_manager = DBManager()
maestras_manager = MaestrasManager()
pdf_generator = PDFGenerator()
auth_manager = AuthManager()

# ========== HELPER FUNCTIONS ==========

def convertir_a_cero(valor):
    """Convertir valores vacíos, None o strings vacíos a 0 para campos numéricos"""
    if valor is None or valor == '' or valor == 'null':
        return 0
    try:
        return int(valor) if valor else 0
    except (ValueError, TypeError):
        return 0

def cargar_query(nombre_archivo):
    """Cargar query SQL desde archivo en carpeta queries/"""
    try:
        ruta_query = os.path.join(os.path.dirname(__file__), 'queries', nombre_archivo)
        with open(ruta_query, 'r', encoding='utf-8') as f:
            # Leer todo el contenido
            contenido = f.read()
            # Eliminar comentarios SQL (líneas que empiezan con --)
            lineas = [linea for linea in contenido.split('\n') 
                     if not linea.strip().startswith('--')]
            # Unir y limpiar espacios extras
            query = '\n'.join(lineas).strip()
            return query
    except FileNotFoundError:
        raise FileNotFoundError(f"No se encontró el archivo de query: {nombre_archivo}")
    except Exception as e:
        raise Exception(f"Error al cargar query {nombre_archivo}: {str(e)}")

# ========== RUTAS DE AUTENTICACIÓN ==========

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        user = auth_manager.verify_user(username, password)
        
        if user:
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['nombre_completo'] = user['nombre_completo']
            flash(f'Bienvenido, {user["nombre_completo"]}!', 'success')
            return redirect(url_for('index'))
        else:
            flash('Usuario o contraseña incorrectos', 'error')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('Sesión cerrada exitosamente', 'success')
    return redirect(url_for('login'))

# ========== GESTIÓN DE USUARIOS (SOLO ADMIN) ==========

@app.route('/gestionar-usuarios')
@login_required
def gestionar_usuarios():
    # Solo admin puede acceder
    if session.get('username') != 'admin':
        flash('No tienes permisos para acceder a esta sección', 'error')
        return redirect(url_for('index'))
    return render_template('gestionar_usuarios.html')

@app.route('/api/usuarios')
@login_required
def api_listar_usuarios():
    if session.get('username') != 'admin':
        return jsonify({'success': False, 'message': 'Sin permisos'}), 403
    
    usuarios = auth_manager.get_all_users()
    return jsonify({'success': True, 'usuarios': usuarios})

@app.route('/api/crear-usuario', methods=['POST'])
@login_required
def api_crear_usuario():
    if session.get('username') != 'admin':
        return jsonify({'success': False, 'message': 'Sin permisos'}), 403
    
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')
    nombre_completo = data.get('nombre_completo')
    email = data.get('email')
    
    # Crear usuario con la contraseña proporcionada
    success, result = auth_manager.create_user(username, password, nombre_completo, email)
    
    if success:
        return jsonify({'success': True, 'message': 'Usuario creado exitosamente', 'user_id': result})
    else:
        return jsonify({'success': False, 'message': result})

@app.route('/api/cambiar-password', methods=['POST'])
@login_required
def api_cambiar_password():
    if session.get('username') != 'admin':
        return jsonify({'success': False, 'message': 'Sin permisos'}), 403
    
    data = request.get_json()
    user_id = data.get('user_id')
    nueva_password = data.get('nueva_password')
    
    if not nueva_password or len(nueva_password) < 4:
        return jsonify({'success': False, 'message': 'La contraseña debe tener al menos 4 caracteres'})
    
    success, message = auth_manager.change_user_password(user_id, nueva_password)
    
    if success:
        return jsonify({'success': True, 'message': message})
    else:
        return jsonify({'success': False, 'message': message})

@app.route('/api/toggle-usuario', methods=['POST'])
@login_required
def api_toggle_usuario():
    if session.get('username') != 'admin':
        return jsonify({'success': False, 'message': 'Sin permisos'}), 403
    
    data = request.get_json()
    user_id = data.get('user_id')
    
    success = auth_manager.toggle_user_status(user_id)
    
    if success:
        return jsonify({'success': True, 'message': 'Estado actualizado'})
    else:
        return jsonify({'success': False, 'message': 'Error al actualizar'})

@app.route('/api/eliminar-usuario', methods=['POST'])
@login_required
def api_eliminar_usuario():
    if session.get('username') != 'admin':
        return jsonify({'success': False, 'message': 'Sin permisos'}), 403
    
    data = request.get_json()
    user_id = data.get('user_id')
    
    success, message = auth_manager.delete_user(user_id)
    
    if success:
        return jsonify({'success': True, 'message': message})
    else:
        return jsonify({'success': False, 'message': message})

# ========== RUTAS PRINCIPALES (PROTEGIDAS) ==========

@app.route('/')
@login_required
def index():
    return render_template('index.html')

@app.route('/nuevo-viaje')
@login_required
def nuevo_viaje():
    casinos = maestras_manager.obtener_todos_casinos()
    choferes_data = maestras_manager.obtener_todos_los_choferes()
    choferes = [c['nombre'] for c in choferes_data]
    centros_costo = maestras_manager.obtener_todos_centros_costo()
    conn = db_manager.get_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT DISTINCT patente_camion FROM viajes WHERE patente_camion IS NOT NULL ORDER BY patente_camion')
    patentes = [row[0] for row in cursor.fetchall()]
    conn.close()
    return render_template('nuevo_viaje.html', casinos=casinos, choferes=choferes, patentes=patentes, centros_costo=centros_costo)

@app.route('/api/buscar-centros-costo/<numero_viaje>')
@login_required
def buscar_centros_costo(numero_viaje):
    """Devuelve lista de centros con su información de casino"""
    try:
        print(f"\n>>> Buscando centros para viaje: {numero_viaje}")
        centros_codigos = db_manager.get_centros_costo_por_viaje(numero_viaje)
        print(f">>> Centros encontrados: {centros_codigos}")
        centros_info = []
        
        for codigo in centros_codigos:
            casino_data = maestras_manager.buscar_casino_por_codigo(int(codigo))
            centros_info.append({
                'codigo': codigo,
                'casino': casino_data['casino'] if casino_data else 'Sin casino',
                'ruta': casino_data['ruta'] if casino_data else ''
            })
        
        print(f">>> Retornando {len(centros_info)} centros")
        return jsonify(centros_info)
    except Exception as e:
        print(f">>> ERROR en buscar_centros_costo: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/obtener-centros-costo')
@login_required
def obtener_centros_costo():
    """Endpoint para obtener todos los centros de costo"""
    centros = maestras_manager.obtener_todos_centros_costo()
    return jsonify(centros)

@app.route('/api/centro-costo-detalles/<int:codigo>')
@login_required
def centro_costo_detalles(codigo):
    """Endpoint que retorna casino y ruta para un código de centro de costo"""
    try:
        centro = maestras_manager.buscar_casino_por_codigo(codigo)
        if centro:
            return jsonify({
                'success': True,
                'casino': centro['casino'],
                'ruta': centro['ruta']
            })
        else:
            return jsonify({'success': False, 'message': 'Centro de costo no encontrado'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/obtener-choferes-completo')
@login_required
def obtener_choferes_completo():
    choferes = maestras_manager.obtener_todos_los_choferes()
    return jsonify(choferes)

@app.route('/test-cargar')
def test_cargar():
    return render_template('test_cargar.html')

@app.route('/api/buscar-viaje-numero/<numero_viaje>')
@login_required
def buscar_viaje_numero(numero_viaje):
    try:
        print(f"\nBUSCANDO VIAJE: {numero_viaje}")
        conn = db_manager.get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT * FROM viajes 
            WHERE numero_viaje = ? 
            ORDER BY id DESC LIMIT 1
        """, (numero_viaje,))
        
        row = cursor.fetchone()
        
        if row:
            columns = [description[0] for description in cursor.description]
            viaje = dict(zip(columns, row))
            
            # Obtener comidas del viaje
            cursor.execute('''
                SELECT id, numero_viaje, numero_centro_costo, guia_comida, 
                       descripcion, kilo, bultos, proveedor 
                FROM comidas_preparadas 
                WHERE numero_viaje = ? AND numero_centro_costo = ?
                ORDER BY id
            ''', (numero_viaje, viaje.get('costo_codigo')))
            
            comidas_rows = cursor.fetchall()
            comidas = []
            for row in comidas_rows:
                comidas.append({
                    'id': row[0],
                    'numero_viaje': row[1],
                    'numero_centro_costo': row[2],
                    'guia_comida': row[3],
                    'descripcion': row[4],
                    'kilo': row[5],
                    'bultos': row[6],
                    'proveedor': row[7]
                })
            
            conn.close()
            
            print(f"VIAJE ENCONTRADO:")
            print(f"   - Conductor: {viaje.get('conductor')}")
            print(f"   - Centro Costo: {viaje.get('costo_codigo')}")
            print(f"   - Salida: {viaje.get('fecha_hora_salida_dhl')}")
            print(f"   - Llegada: {viaje.get('fecha_hora_llegada_dhl')}")
            print(f"   - Celular: {viaje.get('celular')}")
            print(f"   - RUT: {viaje.get('rut')}")
            print(f"   - Comidas: {len(comidas)}")
            return jsonify({'success': True, 'viaje': viaje, 'comidas': comidas})
        else:
            conn.close()
            print(f"VIAJE NO ENCONTRADO")
            return jsonify({'success': False})
    except Exception as e:
        print(f"ERROR: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/casino-por-centro/<centro_costo>')
def casino_por_centro(centro_costo):
    try:
        conn = db_manager.get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT codigo, nombre, ruta 
            FROM casinos 
            WHERE centro_costo = ?
            LIMIT 1
        """, (centro_costo,))
        
        row = cursor.fetchone()
        conn.close()
        
        if row:
            casino = {
                'codigo': row[0],
                'nombre': row[1],
                'ruta': row[2]
            }
            return jsonify({'success': True, 'casino': casino})
        else:
            return jsonify({'success': False})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/actualizar-casino', methods=['POST'])
def actualizar_casino():
    try:
        data = request.json
        conn = db_manager.get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE casinos 
            SET nombre = ?, ruta = ?
            WHERE codigo = ?
        """, (data.get('nombre'), data.get('ruta'), data.get('codigo')))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/actualizar-conductor', methods=['POST'])
def actualizar_conductor():
    try:
        data = request.json
        conn = db_manager.get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE choferes 
            SET celular = ?, rut = ?
            WHERE nombre = ?
        """, (data.get('celular'), data.get('rut'), data.get('nombre')))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/guardar-viaje', methods=['POST'])
def guardar_viaje():
    try:
        data = request.json
        
        # Preparar datos del viaje (convertir a mayúsculas donde corresponda)
        viaje_data = {
            'numero_viaje': data.get('numero_viaje'),
            'costo_codigo': data.get('centro_costo'),
            'fecha': data.get('fecha'),
            'casino': str(data.get('codigo_casino', '')).upper(),
            'ruta': str(data.get('ruta', '')).upper(),
            'tipo_camion': data.get('tipo_camion', ''),
            'patente_camion': str(data.get('patente_camion', '')).upper(),
            'patente_semi': str(data.get('patente_semi', '')).upper(),
            'numero_rampa': data.get('numero_rampa', ''),
            'peso_camion': data.get('peso_camion', ''),
            'numero_camion': data.get('numero_camion', ''),
            'termografos_gps': str(data.get('termografos_gps', '')).upper(),
            'conductor': str(data.get('chofer', '')).upper(),
            'celular': data.get('celular', ''),
            'rut': str(data.get('rut', '')).upper(),
            'fecha_hora_salida_dhl': data.get('hora_salida', ''),
            'fecha_hora_llegada_dhl': data.get('hora_llegada', ''),
            'num_wencos': convertir_a_cero(data.get('num_wencos')),
            'bin': convertir_a_cero(data.get('bin')),
            'pallets': convertir_a_cero(data.get('pallets')),
            'pallets_chep': convertir_a_cero(data.get('pallets_chep')),
            'pallets_pl_negro_grueso': convertir_a_cero(data.get('pallets_pl_negro_grueso')),
            'pallets_pl_negro_alternativo': convertir_a_cero(data.get('pallets_pl_negro_alternativo')),
            'pallets_congelado': convertir_a_cero(data.get('pallets_congelado')),
            'wencos_congelado': convertir_a_cero(data.get('wencos_congelado')),
            'check_congelado': data.get('check_congelado', ''),
            'pallets_refrigerado': convertir_a_cero(data.get('pallets_refrigerado')),
            'wencos_refrigerado': convertir_a_cero(data.get('wencos_refrigerado')),
            'check_refrigerado': data.get('check_refrigerado', ''),
            'pallets_abarrote': convertir_a_cero(data.get('pallets_abarrote')),

            'check_abarrote': data.get('check_abarrote', ''),
            'check_implementos': data.get('check_implementos', ''),
            'check_aseo': data.get('check_aseo', ''),
            'check_trazabilidad': data.get('check_trazabilidad', ''),
            'check_plataforma_wtck': data.get('check_plataforma_wtck', ''),
            'check_env_correo_wtck': data.get('check_env_correo_wtck', ''),
            'check_revision_planilla_despacho': data.get('check_revision_planilla_despacho', ''),
            'guia_1': str(data.get('guia_1', '')).upper(),
            'guia_2': str(data.get('guia_2', '')).upper(),
            'guia_3': str(data.get('guia_3', '')).upper(),
            'guia_4': str(data.get('guia_4', '')).upper(),
            'guia_5': str(data.get('guia_5', '')).upper(),
            'guia_6': str(data.get('guia_6', '')).upper(),
            'guia_7': str(data.get('guia_7', '')).upper(),
            'guia_8': str(data.get('guia_8', '')).upper(),
            'guia_9': str(data.get('guia_9', '')).upper(),
            'guia_10': str(data.get('guia_10', '')).upper(),
            'guia_11': str(data.get('guia_11', '')).upper(),
            'guia_12': str(data.get('guia_12', '')).upper(),
            'guia_13': str(data.get('guia_13', '')).upper(),
            'guia_14': str(data.get('guia_14', '')).upper(),
            'guia_15': str(data.get('guia_15', '')).upper(),
            'guia_16': str(data.get('guia_16', '')).upper(),
            'guia_17': str(data.get('guia_17', '')).upper(),
            'guia_18': str(data.get('guia_18', '')).upper(),
            'guia_19': str(data.get('guia_19', '')).upper(),
            'guia_20': str(data.get('guia_20', '')).upper(),
            'guia_21': str(data.get('guia_21', '')).upper(),
            'sello_salida_1p': str(data.get('sello_salida_1p', '')).upper(),
            'sello_salida_2p': str(data.get('sello_salida_2p', '')).upper(),
            'sello_salida_3p': str(data.get('sello_salida_3p', '')).upper(),
            'sello_salida_4p': str(data.get('sello_salida_4p', '')).upper(),
            'sello_salida_5p': str(data.get('sello_salida_5p', '')).upper(),
            'sello_retorno_1p': str(data.get('sello_retorno_1p', '')).upper(),
            'sello_retorno_2p': str(data.get('sello_retorno_2p', '')).upper(),
            'sello_retorno_3p': str(data.get('sello_retorno_3p', '')).upper(),
            'sello_retorno_4p': str(data.get('sello_retorno_4p', '')).upper(),
            'sello_retorno_5p': str(data.get('sello_retorno_5p', '')).upper(),
            'numero_certificado_fumigacion': str(data.get('numero_certificado_fumigacion', '')).upper(),
            'revision_limpieza_camion_acciones': str(data.get('revision_limpieza_camion_acciones', '')).upper(),
            'administrativo_responsable': str(data.get('administrativo_responsable', '')).upper()
        }
        
        # Insertar viaje
        result = db_manager.insert_viaje(viaje_data)
        
        if result:
            # Insertar comidas preparadas con el centro de costo del viaje
            comidas = data.get('comidas', [])
            if comidas:
                conn = db_manager.get_connection()
                cursor = conn.cursor()
                
                for comida in comidas:
                    cursor.execute("""
                        INSERT INTO comidas_preparadas (
                            numero_viaje, numero_centro_costo, guia_comida, 
                            descripcion, kilo, bultos, proveedor
                        ) VALUES (?, ?, ?, ?, ?, ?, ?)
                    """, (
                        data.get('numero_viaje'),
                        data.get('centro_costo'),
                        str(comida.get('guia_comida', '')).upper(),
                        str(comida.get('descripcion', '')).upper(),
                        convertir_a_cero(comida.get('kilo')),
                        convertir_a_cero(comida.get('bultos')),
                        str(comida.get('proveedor', '')).upper()
                    ))
                
                conn.commit()
                conn.close()
            
            mensaje = f'Viaje {data.get("numero_viaje")} creado exitosamente'
            if comidas:
                mensaje += f' con {len(comidas)} comida(s) preparada(s)'
            
            return jsonify({'success': True, 'message': mensaje})
        else:
            return jsonify({'success': False, 'message': 'Error al guardar el viaje'}), 500
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/editar-viaje')
@login_required
def editar_viaje():
    return render_template('editar_viaje.html')

@app.route('/api/buscar-viaje', methods=['POST'])
@login_required
def buscar_viaje():
    try:
        data = request.json
        numero_viaje = data.get('numero_viaje')
        centro_costo = data.get('centro_costo')
        
        # Obtener viaje
        conn = db_manager.get_connection()
        cursor = conn.cursor()
        
        cursor.execute('SELECT * FROM viajes WHERE numero_viaje = ? AND costo_codigo = ?', 
                      (numero_viaje, centro_costo))
        viaje_row = cursor.fetchone()
        
        if not viaje_row:
            conn.close()
            return jsonify({'success': False, 'message': 'Viaje no encontrado'}), 404
        
        # Convertir a diccionario
        columns = [description[0] for description in cursor.description]
        viaje = dict(zip(columns, viaje_row))
        
        # Obtener comidas
        cursor.execute('''
            SELECT id, numero_viaje, numero_centro_costo, guia_comida, 
                   descripcion, kilo, bultos, proveedor 
            FROM comidas_preparadas 
            WHERE numero_viaje = ? AND numero_centro_costo = ?
            ORDER BY id
        ''', (numero_viaje, centro_costo))
        
        comidas_rows = cursor.fetchall()
        comidas = []
        for row in comidas_rows:
            comidas.append({
                'id': row[0],
                'numero_viaje': row[1],
                'numero_centro_costo': row[2],
                'guia_comida': row[3],
                'descripcion': row[4],
                'kilo': row[5],
                'bultos': row[6],
                'proveedor': row[7]
            })
        
        conn.close()
        
        return jsonify({
            'success': True, 
            'viaje': viaje, 
            'comidas': comidas
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/actualizar-viaje', methods=['POST'])
def actualizar_viaje():
    try:
        data = request.json
        viaje_data = {
            'fecha': data.get('fecha'),
            'casino': str(data.get('casino', '')).upper(),
            'ruta': str(data.get('ruta', '')).upper(),
            'tipo_camion': data.get('tipo_camion'),
            'patente_camion': str(data.get('patente_camion', '')).upper(),
            'patente_semi': str(data.get('patente_semi', '')).upper(),
            'numero_rampa': data.get('numero_rampa'),
            'peso_camion': data.get('peso_camion'),
            'numero_camion': data.get('numero_camion'),
            'termografos_gps': str(data.get('termografos_gps', '')).upper(),
            'conductor': str(data.get('chofer', '')).upper(),
            'celular': data.get('celular'),
            'rut': str(data.get('rut', '')).upper(),
            'fecha_hora_salida_dhl': data.get('hora_salida'),
            'fecha_hora_llegada_dhl': data.get('hora_llegada'),
            'num_wencos': convertir_a_cero(data.get('num_wencos')),
            'bin': convertir_a_cero(data.get('bin')),
            'pallets': convertir_a_cero(data.get('pallets')),
            'pallets_chep': convertir_a_cero(data.get('pallets_chep')),
            'pallets_pl_negro_grueso': convertir_a_cero(data.get('pallets_pl_negro_grueso')),
            'pallets_pl_negro_alternativo': convertir_a_cero(data.get('pallets_pl_negro_alternativo')),
            'pallets_congelado': convertir_a_cero(data.get('pallets_congelado')),
            'wencos_congelado': convertir_a_cero(data.get('wencos_congelado')),
            'check_congelado': data.get('check_congelado', ''),
            'pallets_refrigerado': convertir_a_cero(data.get('pallets_refrigerado')),
            'wencos_refrigerado': convertir_a_cero(data.get('wencos_refrigerado')),
            'check_refrigerado': data.get('check_refrigerado', ''),
            'pallets_abarrote': convertir_a_cero(data.get('pallets_abarrote')),

            'check_abarrote': data.get('check_abarrote', ''),
            'check_implementos': data.get('check_implementos', ''),
            'check_aseo': data.get('check_aseo', ''),
            'check_trazabilidad': data.get('check_trazabilidad', ''),
            'check_plataforma_wtck': data.get('check_plataforma_wtck', ''),
            'check_env_correo_wtck': data.get('check_env_correo_wtck', ''),
            'check_revision_planilla_despacho': data.get('check_revision_planilla_despacho', ''),
            'guia_1': str(data.get('guia_1', '')).upper(),
            'guia_2': str(data.get('guia_2', '')).upper(),
            'guia_3': str(data.get('guia_3', '')).upper(),
            'guia_4': str(data.get('guia_4', '')).upper(),
            'guia_5': str(data.get('guia_5', '')).upper(),
            'guia_6': str(data.get('guia_6', '')).upper(),
            'guia_7': str(data.get('guia_7', '')).upper(),
            'guia_8': str(data.get('guia_8', '')).upper(),
            'guia_9': str(data.get('guia_9', '')).upper(),
            'guia_10': str(data.get('guia_10', '')).upper(),
            'guia_11': str(data.get('guia_11', '')).upper(),
            'guia_12': str(data.get('guia_12', '')).upper(),
            'guia_13': str(data.get('guia_13', '')).upper(),
            'guia_14': str(data.get('guia_14', '')).upper(),
            'guia_15': str(data.get('guia_15', '')).upper(),
            'guia_16': str(data.get('guia_16', '')).upper(),
            'guia_17': str(data.get('guia_17', '')).upper(),
            'guia_18': str(data.get('guia_18', '')).upper(),
            'guia_19': str(data.get('guia_19', '')).upper(),
            'guia_20': str(data.get('guia_20', '')).upper(),
            'guia_21': str(data.get('guia_21', '')).upper(),
            'sello_salida_1p': str(data.get('sello_salida_1p', '')).upper(),
            'sello_salida_2p': str(data.get('sello_salida_2p', '')).upper(),
            'sello_salida_3p': str(data.get('sello_salida_3p', '')).upper(),
            'sello_salida_4p': str(data.get('sello_salida_4p', '')).upper(),
            'sello_salida_5p': str(data.get('sello_salida_5p', '')).upper(),
            'sello_retorno_1p': str(data.get('sello_retorno_1p', '')).upper(),
            'sello_retorno_2p': str(data.get('sello_retorno_2p', '')).upper(),
            'sello_retorno_3p': str(data.get('sello_retorno_3p', '')).upper(),
            'sello_retorno_4p': str(data.get('sello_retorno_4p', '')).upper(),
            'sello_retorno_5p': str(data.get('sello_retorno_5p', '')).upper(),
            'numero_certificado_fumigacion': str(data.get('numero_certificado_fumigacion', '')).upper(),
            'revision_limpieza_camion_acciones': str(data.get('revision_limpieza_camion_acciones', '')).upper(),
            'administrativo_responsable': str(data.get('administrativo_responsable', '')).upper()
        }
        
        # Actualizar viaje
        db_manager.update_viaje_by_numero_centro(data.get('numero_viaje'), data.get('centro_costo'), viaje_data)
        
        # Actualizar comidas - eliminar las existentes e insertar las nuevas
        db_manager.delete_comidas_by_viaje_centro(data.get('numero_viaje'), data.get('centro_costo'))
        
        comidas = data.get('comidas', [])
        if comidas:
            conn = db_manager.get_connection()
            cursor = conn.cursor()
            
            for comida in comidas:
                cursor.execute("""
                    INSERT INTO comidas_preparadas (
                        numero_viaje, numero_centro_costo, guia_comida, 
                        descripcion, kilo, bultos, proveedor
                    ) VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (
                    data.get('numero_viaje'),
                    data.get('centro_costo'),
                    str(comida.get('guia_comida', '')).upper(),
                    str(comida.get('descripcion', '')).upper(),
                    convertir_a_cero(comida.get('kilo')),
                    convertir_a_cero(comida.get('bultos')),
                    str(comida.get('proveedor', '')).upper()
                ))
            
            conn.commit()
            conn.close()
        
        return jsonify({'success': True, 'message': 'Viaje actualizado correctamente'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/eliminar-viaje', methods=['POST'])
def eliminar_viaje():
    try:
        data = request.json
        numero_viaje = data.get('numero_viaje')
        centro_costo = data.get('centro_costo')
        
        print(f"\n[ELIMINAR_VIAJE] Solicitud recibida: numero_viaje={numero_viaje}, centro_costo={centro_costo}")
        
        # Buscar el viaje
        viaje = db_manager.get_viaje_por_numero_y_centro(numero_viaje, centro_costo)
        
        if viaje:
            print(f"[ELIMINAR_VIAJE] Viaje encontrado con ID: {viaje['id']}")
            
            # Eliminar comidas asociadas primero
            db_manager.delete_comidas_by_viaje_centro(numero_viaje, centro_costo)
            
            # Eliminar el viaje por su ID
            db_manager.delete_viaje(viaje['id'])
            
            print(f"[ELIMINAR_VIAJE] Viaje {numero_viaje} eliminado exitosamente\n")
            return jsonify({'success': True, 'message': f'Viaje {numero_viaje} eliminado correctamente'})
        else:
            print(f"[ELIMINAR_VIAJE] ERROR: Viaje no encontrado\n")
            return jsonify({'success': False, 'message': 'Viaje no encontrado'}), 404
    except Exception as e:
        print(f"[ELIMINAR_VIAJE] EXCEPCIÓN: {str(e)}\n")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/generar-pdf')
@login_required
def generar_pdf_page():
    return render_template('generar_pdf.html')

@app.route('/dashboard')
@login_required
def dashboard():
    return render_template('dashboard.html')

@app.route('/reportes')
@login_required
def reportes():
    return render_template('reportes.html')

@app.route('/api/maestras/administrativos')
def get_administrativos():
    """API para obtener lista de administrativos"""
    try:
        conn = db_manager.get_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT nombre FROM maestras_administrativos ORDER BY nombre')
        administrativos = [{'nombre': row[0]} for row in cursor.fetchall()]
        conn.close()
        return jsonify(administrativos)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/dashboard/estadisticas')
@login_required
def get_estadisticas_dashboard():
    """API para obtener estadísticas del dashboard"""
    try:
        from datetime import datetime, timedelta
        
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        administrativo = request.args.get('administrativo', '')
        vista = request.args.get('vista', 'diaria')
        
        conn = db_manager.get_connection()
        cursor = conn.cursor()
        
        # Query base
        where_clause = "WHERE fecha BETWEEN ? AND ?"
        params = [fecha_inicio, fecha_fin]
        
        if administrativo:
            where_clause += " AND administrativo_responsable = ?"
            params.append(administrativo)
        
        # Resumen general - Solo activos y validaciones requeridas
        cursor.execute(f'''
            SELECT 
                COUNT(*) as total_viajes,
                SUM(CASE WHEN check_congelado = 'X' THEN 1 ELSE 0 END) as check_congelado,
                SUM(CASE WHEN check_refrigerado = 'X' THEN 1 ELSE 0 END) as check_refrigerado,
                SUM(CASE WHEN check_abarrote = 'X' THEN 1 ELSE 0 END) as check_abarrote,
                SUM(CASE WHEN check_implementos = 'X' THEN 1 ELSE 0 END) as check_implementos,
                SUM(CASE WHEN check_aseo = 'X' THEN 1 ELSE 0 END) as check_aseo,
                SUM(CASE WHEN check_trazabilidad = 'X' THEN 1 ELSE 0 END) as check_trazabilidad,
                COALESCE(SUM(CAST(COALESCE(num_wencos, '0') AS INTEGER)), 0) as total_wencos,
                COALESCE(SUM(CAST(COALESCE(bin, '0') AS INTEGER)), 0) as total_bin,
                COALESCE(SUM(CAST(COALESCE(pallets, '0') AS INTEGER)), 0) as pallets_std,
                COALESCE(SUM(CAST(COALESCE(pallets_chep, '0') AS INTEGER)), 0) as pallets_chep,
                COALESCE(SUM(CAST(COALESCE(pallets_pl_negro_grueso, '0') AS INTEGER)), 0) as pallets_negro_grueso,
                COALESCE(SUM(CAST(COALESCE(pallets_pl_negro_alternativo, '0') AS INTEGER)), 0) as pallets_negro_alternativo
            FROM viajes
            {where_clause}
        ''', params)
        
        resumen = cursor.fetchone()
        
        # Calcular total de pallets
        total_pallets = (resumen[9] or 0) + (resumen[10] or 0) + (resumen[11] or 0) + (resumen[12] or 0)
        
        # Tendencia temporal según vista
        if vista == 'diaria':
            group_format = '%Y-%m-%d'
        elif vista == 'semanal':
            group_format = '%Y-W%W'
        else:  # mensual
            group_format = '%Y-%m'
        
        cursor.execute(f'''
            SELECT 
                strftime('{group_format}', fecha) as periodo,
                COUNT(*) as total
            FROM viajes
            {where_clause}
            GROUP BY periodo
            ORDER BY periodo
        ''', params)
        
        tendencia = [{'periodo': row[0], 'total': row[1]} for row in cursor.fetchall()]
        
        # Tendencia por tipos - incluir las 6 validaciones
        cursor.execute(f'''
            SELECT 
                strftime('{group_format}', fecha) as periodo,
                SUM(CASE WHEN check_congelado = 'X' THEN 1 ELSE 0 END) as congelado,
                SUM(CASE WHEN check_refrigerado = 'X' THEN 1 ELSE 0 END) as refrigerado,
                SUM(CASE WHEN check_abarrote = 'X' THEN 1 ELSE 0 END) as abarrote,
                SUM(CASE WHEN check_implementos = 'X' THEN 1 ELSE 0 END) as implementos,
                SUM(CASE WHEN check_aseo = 'X' THEN 1 ELSE 0 END) as aseo,
                SUM(CASE WHEN check_trazabilidad = 'X' THEN 1 ELSE 0 END) as trazabilidad
            FROM viajes
            {where_clause}
            GROUP BY periodo
            ORDER BY periodo
        ''', params)
        
        tendencia_tipos = [{'periodo': row[0], 'congelado': row[1], 'refrigerado': row[2], 'abarrote': row[3], 'implementos': row[4], 'aseo': row[5], 'trazabilidad': row[6]} for row in cursor.fetchall()]
        
        # Por administrativo
        cursor.execute(f'''
            SELECT 
                COALESCE(administrativo_responsable, 'Sin asignar') as administrativo,
                COUNT(*) as total
            FROM viajes
            {where_clause}
            GROUP BY administrativo_responsable
            ORDER BY total DESC
        ''', params)
        
        por_administrativo = [{'administrativo': row[0], 'total': row[1]} for row in cursor.fetchall()]
        
        # Top 10 casinos
        cursor.execute(f'''
            SELECT 
                COALESCE(casino, 'Sin especificar') as casino,
                COUNT(*) as total
            FROM viajes
            {where_clause}
            GROUP BY casino
            ORDER BY total DESC
            LIMIT 10
        ''', params)
        
        top_casinos = [{'casino': row[0], 'total': row[1]} for row in cursor.fetchall()]
        
        # Detalle para tabla con pallets y wencos totales
        cursor.execute(f'''
            SELECT 
                strftime('{group_format}', fecha) as periodo,
                COALESCE(administrativo_responsable, 'Sin asignar') as administrativo,
                COUNT(*) as total,
                SUM(CASE WHEN check_congelado = 'X' THEN 1 ELSE 0 END) as congelado,
                SUM(CASE WHEN check_refrigerado = 'X' THEN 1 ELSE 0 END) as refrigerado,
                SUM(CASE WHEN check_abarrote = 'X' THEN 1 ELSE 0 END) as abarrote,
                SUM(CASE WHEN check_implementos = 'X' THEN 1 ELSE 0 END) as implementos,
                SUM(CASE WHEN check_aseo = 'X' THEN 1 ELSE 0 END) as aseo,
                SUM(CASE WHEN check_trazabilidad = 'X' THEN 1 ELSE 0 END) as trazabilidad,
                COALESCE(SUM(
                    CAST(COALESCE(pallets, '0') AS INTEGER) +
                    CAST(COALESCE(pallets_chep, '0') AS INTEGER) +
                    CAST(COALESCE(pallets_pl_negro_grueso, '0') AS INTEGER) +
                    CAST(COALESCE(pallets_pl_negro_alternativo, '0') AS INTEGER)
                ), 0) as pallets,
                COALESCE(SUM(CAST(COALESCE(num_wencos, '0') AS INTEGER)), 0) as wencos
            FROM viajes
            {where_clause}
            GROUP BY periodo, administrativo_responsable
            ORDER BY periodo DESC, total DESC
        ''', params)
        
        detalle = [{
            'periodo': row[0],
            'administrativo': row[1],
            'total': row[2],
            'congelado': row[3],
            'refrigerado': row[4],
            'abarrote': row[5],
            'implementos': row[6],
            'aseo': row[7],
            'trazabilidad': row[8],
            'pallets': row[9],
            'wencos': row[10]
        } for row in cursor.fetchall()]
        
        conn.close()
        
        return jsonify({
            'resumen': {
                'total_viajes': resumen[0],
                'check_congelado': resumen[1],
                'check_refrigerado': resumen[2],
                'check_abarrote': resumen[3],
                'check_implementos': resumen[4],
                'check_aseo': resumen[5],
                'check_trazabilidad': resumen[6],
                'total_wencos': resumen[7],
                'total_bin': resumen[8],
                'pallets_std': resumen[9],
                'pallets_chep': resumen[10],
                'pallets_negro_grueso': resumen[11],
                'pallets_negro_alternativo': resumen[12],
                'total_pallets': total_pallets
            },
            'tendencia': tendencia,
            'tendencia_tipos': tendencia_tipos,
            'por_administrativo': por_administrativo,
            'top_casinos': top_casinos,
            'detalle': detalle
        })
        
    except Exception as e:
        print(f"Error en dashboard: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/descargar-reporte-casinos')
@login_required
def descargar_reporte_casinos():
    """Generar y descargar reporte Excel de la maestra de casinos"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        from datetime import datetime
        
        # Cargar query desde archivo
        query = cargar_query('reporte_maestra_casinos.sql')
        
        # Ejecutar consulta
        conn = db_manager.get_connection()
        cursor = conn.cursor()
        cursor.execute(query)
        casinos_data = cursor.fetchall()
        conn.close()
        
        # Crear libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Maestra Casinos"
        
        # Estilos
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Encabezados
        headers = ['CÓDIGO CENTRO COSTO', 'CASINO', 'RUTA']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Datos
        for row_num, casino in enumerate(casinos_data, 2):
            ws.cell(row=row_num, column=1, value=casino[0])  # codigo_costo
            ws.cell(row=row_num, column=2, value=casino[1])  # casino
            ws.cell(row=row_num, column=3, value=casino[2])  # ruta
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 30
        
        # Guardar en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Nombre del archivo con fecha
        fecha = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'Maestra_Casinos_{fecha}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/descargar-reporte-choferes')
@login_required
def descargar_reporte_choferes():
    """Generar y descargar reporte Excel de la maestra de choferes"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        from datetime import datetime
        
        # Cargar query desde archivo
        query = cargar_query('reporte_maestra_choferes.sql')
        
        # Ejecutar consulta
        conn = db_manager.get_connection()
        cursor = conn.cursor()
        cursor.execute(query)
        choferes_data = cursor.fetchall()
        conn.close()
        
        # Crear libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Maestra Choferes"
        
        # Estilos
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Encabezados
        headers = ['NOMBRE', 'RUT', 'CELULAR']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Datos
        for row_num, chofer in enumerate(choferes_data, 2):
            ws.cell(row=row_num, column=1, value=chofer[0])  # nombre
            ws.cell(row=row_num, column=2, value=chofer[1])  # rut
            ws.cell(row=row_num, column=3, value=chofer[2])  # celular
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        
        # Guardar en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Nombre del archivo con fecha
        fecha = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'Maestra_Choferes_{fecha}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/descargar-reporte-comidas')
@login_required
def descargar_reporte_comidas():
    """Generar y descargar reporte Excel de comidas e implementos por rango de fechas"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        from datetime import datetime
        
        # Obtener parámetros de fechas
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        
        if not fecha_inicio or not fecha_fin:
            return jsonify({'success': False, 'message': 'Debe proporcionar fecha de inicio y fin'}), 400
        
        # Cargar query desde archivo
        query = cargar_query('reporte_comidas_implementos.sql')
        
        # Ejecutar consulta
        conn = db_manager.get_connection()
        cursor = conn.cursor()
        cursor.execute(query, (fecha_inicio, fecha_fin))
        comidas = cursor.fetchall()
        conn.close()
        
        # Crear libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Comidas e Implementos"
        
        # Estilos
        header_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Encabezados
        headers = ['FECHA', 'NRO VIAJE', 'CASINO', 'CONDUCTOR', 'CENTRO COSTO', 
                   'GUÍA COMIDA', 'DESCRIPCIÓN', 'KILOS', 'BULTOS', 'PROVEEDOR']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Datos
        for row_num, comida in enumerate(comidas, 2):
            ws.cell(row=row_num, column=1, value=comida[0])  # fecha
            ws.cell(row=row_num, column=2, value=comida[1])  # numero_viaje
            ws.cell(row=row_num, column=3, value=comida[2])  # casino
            ws.cell(row=row_num, column=4, value=comida[3])  # conductor
            ws.cell(row=row_num, column=5, value=comida[4])  # numero_centro_costo
            ws.cell(row=row_num, column=6, value=comida[5])  # guia_comida
            ws.cell(row=row_num, column=7, value=comida[6])  # descripcion
            ws.cell(row=row_num, column=8, value=comida[7])  # kilo
            ws.cell(row=row_num, column=9, value=comida[8])  # bultos
            ws.cell(row=row_num, column=10, value=comida[9]) # proveedor
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 12  # Fecha
        ws.column_dimensions['B'].width = 15  # Nro Viaje
        ws.column_dimensions['C'].width = 30  # Casino
        ws.column_dimensions['D'].width = 30  # Conductor
        ws.column_dimensions['E'].width = 15  # Centro Costo
        ws.column_dimensions['F'].width = 15  # Guía Comida
        ws.column_dimensions['G'].width = 40  # Descripción
        ws.column_dimensions['H'].width = 10  # Kilos
        ws.column_dimensions['I'].width = 10  # Bultos
        ws.column_dimensions['J'].width = 25  # Proveedor
        
        # Guardar en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Nombre del archivo con fechas y timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'Comidas_Implementos_{fecha_inicio}_al_{fecha_fin}_{timestamp}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/descargar-reporte-viajes')
@login_required
def descargar_reporte_viajes():
    """Generar y descargar reporte Excel de viajes completos por rango de fechas"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        from datetime import datetime
        
        # Obtener parámetros de fechas
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        
        if not fecha_inicio or not fecha_fin:
            return jsonify({'success': False, 'message': 'Debe proporcionar fecha de inicio y fin'}), 400
        
        # Cargar query desde archivo
        query = cargar_query('reporte_viajes_completos.sql')
        
        # Ejecutar consulta
        conn = db_manager.get_connection()
        cursor = conn.cursor()
        cursor.execute(query, (fecha_inicio, fecha_fin))
        viajes_data = cursor.fetchall()
        conn.close()
        
        # Crear libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Viajes Completos"
        
        # Estilos
        header_fill = PatternFill(start_color="9B59B6", end_color="9B59B6", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Encabezados (todos los campos de viajes)
        headers = [
            'NRO VIAJE', 'CASINO', 'RUTA', 'TIPO CAMIÓN', 'PATENTE CAMIÓN', 'PATENTE SEMI',
            'NRO RAMPA', 'PESO CAMIÓN', 'CÓDIGO COSTO', 'TERMÓGRAFOS GPS', 'FECHA',
            'LLEGADA DHL', 'SALIDA DHL', 'CONDUCTOR', 'CELULAR', 'RUT', 'NRO CAMIÓN',
            'WENCOS', 'BIN', 'PALLETS', 'PALLETS CHEP', 'PALLETS NEGRO GRUESO',
            'PALLETS NEGRO ALT', 'PALLETS REFRIG', 'WENCOS REFRIG', 'PALLETS CONG',
            'WENCOS CONG', 'PALLETS ABARROTE', 'CHECK CONG', 'CHECK REFRIG', 'CHECK ABARROTE',
            'CHECK IMPLEMENTOS', 'CHECK ASEO', 'CHECK TRAZAB', 'CHECK WTCK', 'CHECK CORREO WTCK',
            'CHECK PLANILLA', 'SELLO SAL 1P', 'SELLO SAL 2P', 'SELLO SAL 3P', 'SELLO SAL 4P',
            'SELLO SAL 5P', 'SELLO RET 1P', 'SELLO RET 2P', 'SELLO RET 3P', 'SELLO RET 4P',
            'SELLO RET 5P', 'GUÍA 1', 'GUÍA 2', 'GUÍA 3', 'GUÍA 4', 'GUÍA 5', 'GUÍA 6',
            'GUÍA 7', 'GUÍA 8', 'GUÍA 9', 'GUÍA 10', 'GUÍA 11', 'GUÍA 12', 'GUÍA 13', 'GUÍA 14',
            'GUÍA 15', 'GUÍA 16', 'GUÍA 17', 'GUÍA 18', 'GUÍA 19', 'GUÍA 20', 'GUÍA 21',
            'CERT FUMIGACIÓN', 'REVISIÓN LIMPIEZA', 'ADMIN RESPONSABLE'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Datos
        for row_num, viaje in enumerate(viajes_data, 2):
            for col_num, valor in enumerate(viaje, 1):
                ws.cell(row=row_num, column=col_num, value=valor)
        
        # Ajustar anchos de columna (generales)
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Guardar en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Nombre del archivo con fechas y timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'Viajes_Completos_{fecha_inicio}_al_{fecha_fin}_{timestamp}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/descargar-reporte-facturacion')
@login_required
def descargar_reporte_facturacion():
    """Generar y descargar reporte Excel de facturación por rango de fechas"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        from datetime import datetime
        
        # Obtener parámetros de fechas
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        
        if not fecha_inicio or not fecha_fin:
            return jsonify({'success': False, 'message': 'Debe proporcionar fecha de inicio y fin'}), 400
        
        # Cargar query desde archivo
        query = cargar_query('reporte_facturacion.sql')
        
        # Ejecutar consulta
        conn = db_manager.get_connection()
        cursor = conn.cursor()
        cursor.execute(query, (fecha_inicio, fecha_fin))
        facturacion_data = cursor.fetchall()
        conn.close()
        
        # Crear libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Facturación"
        
        # Estilos
        header_fill = PatternFill(start_color="16A085", end_color="16A085", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Encabezados
        headers = [
            'NRO VIAJE', 'CASINO', 'CÓDIGO COSTO', 'FECHA', 'WENCOS', 'BIN',
            'PALLETS', 'PALLETS CHEP', 'PALLETS NEGRO GRUESO', 'PALLETS NEGRO ALT',
            'GUÍAS'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Datos
        for row_num, factura in enumerate(facturacion_data, 2):
            for col_num, valor in enumerate(factura, 1):
                ws.cell(row=row_num, column=col_num, value=valor)
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 15  # Nro Viaje
        ws.column_dimensions['B'].width = 35  # Casino
        ws.column_dimensions['C'].width = 15  # Código Costo
        ws.column_dimensions['D'].width = 12  # Fecha
        ws.column_dimensions['E'].width = 10  # Wencos
        ws.column_dimensions['F'].width = 10  # Bin
        ws.column_dimensions['G'].width = 10  # Pallets
        ws.column_dimensions['H'].width = 15  # Pallets CHEP
        ws.column_dimensions['I'].width = 20  # Pallets Negro Grueso
        ws.column_dimensions['J'].width = 20  # Pallets Negro Alt
        ws.column_dimensions['K'].width = 60  # Guías
        
        # Guardar en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Nombre del archivo con fechas y timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'Facturacion_{fecha_inicio}_al_{fecha_fin}_{timestamp}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/descargar-reporte-facturacion-diaria')
@login_required
def descargar_reporte_facturacion_diaria():
    """Generar y descargar reporte Excel de control de activos diario"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        from datetime import datetime
        
        # Obtener parámetro de fecha
        fecha = request.args.get('fecha')
        
        if not fecha:
            return jsonify({'success': False, 'message': 'Debe proporcionar la fecha'}), 400
        
        # Cargar query desde archivo
        query = cargar_query('reporte_control_activos_diario.sql')
        
        # Ejecutar consulta
        conn = db_manager.get_connection()
        cursor = conn.cursor()
        cursor.execute(query, (fecha,))
        facturacion_data = cursor.fetchall()
        conn.close()
        
        # Crear libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Control Activos Diario"
        
        # Estilos
        header_fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Encabezados
        headers = [
            'NRO VIAJE', 'CASINO', 'CÓDIGO COSTO', 'FECHA', 'WENCOS', 'BIN',
            'PALLETS', 'PALLETS CHEP', 'PALLETS NEGRO GRUESO', 'PALLETS NEGRO ALT'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Datos
        for row_num, factura in enumerate(facturacion_data, 2):
            for col_num, valor in enumerate(factura, 1):
                ws.cell(row=row_num, column=col_num, value=valor)
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 15  # Nro Viaje
        ws.column_dimensions['B'].width = 35  # Casino
        ws.column_dimensions['C'].width = 15  # Código Costo
        ws.column_dimensions['D'].width = 12  # Fecha
        ws.column_dimensions['E'].width = 10  # Wencos
        ws.column_dimensions['F'].width = 10  # Bin
        ws.column_dimensions['G'].width = 10  # Pallets
        ws.column_dimensions['H'].width = 15  # Pallets CHEP
        ws.column_dimensions['I'].width = 20  # Pallets Negro Grueso
        ws.column_dimensions['J'].width = 20  # Pallets Negro Alt
        
        # Guardar en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Nombre del archivo con fecha y timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'Control_Activos_Diario_{fecha}_{timestamp}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/descargar-reporte-rendiciones')
@login_required
def descargar_reporte_rendiciones():
    """Generar y descargar reporte Excel de rendiciones por rango de fechas"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        from datetime import datetime
        
        # Obtener parámetros
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        
        if not fecha_inicio or not fecha_fin:
            return jsonify({'success': False, 'message': 'Debe proporcionar ambas fechas'}), 400
        
        # Obtener datos
        rendiciones = rendiciones_manager.obtener_rendiciones_por_fecha(fecha_inicio, fecha_fin)
        
        # Crear libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Rendiciones"
        
        # Estilos
        header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Encabezados
        headers = [
            'N° VIAJE', 'PDT', 'RUTA',
            'FECHA CREACIÓN', 'FECHA MODIFICACIÓN', 'ESTADO RENDICIÓN'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Datos
        for row_num, rendicion in enumerate(rendiciones, 2):
            ws.cell(row=row_num, column=1, value=rendicion['nro_viaje'])
            ws.cell(row=row_num, column=2, value=rendicion['pdt'])
            ws.cell(row=row_num, column=3, value=rendicion['ruta'])
            ws.cell(row=row_num, column=4, value=rendicion['fecha_creacion'])
            ws.cell(row=row_num, column=5, value=rendicion['fecha_modificacion'])
            ws.cell(row=row_num, column=6, value=rendicion['estado_rendicion'])
        
        # Ajustar anchos de columna
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        # Guardar en memoria
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Crear respuesta
        response = send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'Rendiciones_{fecha_inicio}_a_{fecha_fin}.xlsx'
        )
        
        return response
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/generar-pdf', methods=['POST'])
def generar_pdf_api():
    try:
        numero_viaje = request.json.get('numero_viaje')
        
        # Obtener todos los centros de costo del viaje
        centros = db_manager.get_centros_costo_por_viaje(numero_viaje)
        if not centros:
            return jsonify({'success': False, 'message': 'No se encontraron centros de costo para este viaje'}), 404
        
        # Generar PDF completo usando la nueva función
        pdf_path = pdf_generator.generar_pdf_completo(numero_viaje)
        
        if pdf_path and os.path.exists(pdf_path):
            num_centros = len(centros)
            return jsonify({
                'success': True, 
                'pdf_filename': os.path.basename(pdf_path),
                'message': f'PDF generado con {num_centros} hoja(s) - una por centro de costo'
            })
        return jsonify({'success': False, 'message': 'Error al generar el PDF'}), 500
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/descargar-pdf/<path:filename>')
@login_required
def descargar_pdf(filename):
    # Determinar el directorio base correcto
    if getattr(sys, 'frozen', False):
        # Corriendo como ejecutable empaquetado
        base_dir = os.path.dirname(sys.executable)
    else:
        # Corriendo como script normal
        base_dir = os.path.dirname(os.path.abspath(__file__))
    
    filepath = os.path.join(base_dir, 'pdfs', filename)
    
    if os.path.exists(filepath):
        return send_file(filepath, as_attachment=True)
    else:
        flash(f'El archivo PDF no existe: {filename}', 'error')
        return redirect(url_for('generar_pdf_page'))

# ====== RUTAS API PARA CREAR REGISTROS EN MAESTRAS ======

@app.route('/api/crear-centro-costo', methods=['POST'])
def crear_centro_costo():
    """Crear un nuevo centro de costo"""
    try:
        data = request.json
        codigo_costo = data.get('codigo_costo', '').strip()
        casino = data.get('casino', '').strip()
        ruta = data.get('ruta', '').strip()
        
        if not codigo_costo or not casino:
            return jsonify({'success': False, 'message': 'Código y Casino son obligatorios'}), 400
        
        resultado = maestras_manager.crear_centro_costo(codigo_costo, casino, ruta)
        
        if resultado:
            return jsonify({'success': True, 'message': 'Centro de Costo creado exitosamente'})
        else:
            return jsonify({'success': False, 'message': 'El centro de costo ya existe'}), 400
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/crear-chofer', methods=['POST'])
def crear_chofer():
    """Crear un nuevo chofer"""
    try:
        data = request.json
        nombre = data.get('nombre', '').strip().upper()
        rut = data.get('rut', '').strip()
        celular = data.get('celular', '').strip()
        
        if not nombre or not rut:
            return jsonify({'success': False, 'message': 'Nombre y RUT son obligatorios'}), 400
        
        resultado = maestras_manager.crear_chofer(nombre, rut, celular)
        
        if resultado:
            return jsonify({'success': True, 'message': 'Chofer creado exitosamente'})
        else:
            return jsonify({'success': False, 'message': 'El chofer ya existe (RUT duplicado)'}), 400
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/crear-administrativo', methods=['POST'])
def crear_administrativo():
    """Crear un nuevo administrativo"""
    try:
        data = request.json
        nombre = data.get('nombre', '').strip().upper()
        
        if not nombre:
            return jsonify({'success': False, 'message': 'Nombre es obligatorio'}), 400
        
        resultado = maestras_manager.crear_administrativo(nombre)
        
        if resultado:
            return jsonify({'success': True, 'message': 'Administrativo creado exitosamente'})
        else:
            return jsonify({'success': False, 'message': 'El administrativo ya existe'}), 400
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/listar-administrativos')
def listar_administrativos():
    """Listar todos los nombres de administrativos"""
    try:
        nombres = maestras_manager.listar_administrativos_nombres()
        return jsonify({'success': True, 'administrativos': nombres})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

# ====== RUTAS API PARA ACTUALIZAR REGISTROS EN MAESTRAS ======

@app.route('/api/actualizar-centro-costo', methods=['POST'])
def actualizar_centro_costo():
    """Actualizar un centro de costo existente"""
    try:
        data = request.json
        codigo_costo = data.get('codigo_costo', '').strip()
        casino = data.get('casino', '').strip()
        ruta = data.get('ruta', '').strip()
        
        if not codigo_costo or not casino:
            return jsonify({'success': False, 'message': 'Código y Casino son obligatorios'}), 400
        
        resultado = maestras_manager.actualizar_centro_costo_por_codigo(codigo_costo, casino, ruta)
        
        if resultado:
            return jsonify({'success': True, 'message': 'Centro de Costo actualizado exitosamente'})
        else:
            return jsonify({'success': False, 'message': 'No se encontró el centro de costo'}), 404
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/actualizar-chofer', methods=['POST'])
def actualizar_chofer():
    """Actualizar un chofer existente"""
    try:
        data = request.json
        nombre = data.get('nombre', '').strip()
        rut = data.get('rut', '').strip()
        celular = data.get('celular', '').strip()
        
        if not nombre:
            return jsonify({'success': False, 'message': 'Nombre es obligatorio'}), 400
        
        resultado = maestras_manager.actualizar_chofer_por_nombre(nombre, rut, celular)
        
        if resultado:
            return jsonify({'success': True, 'message': 'Chofer actualizado exitosamente'})
        else:
            return jsonify({'success': False, 'message': 'No se encontró el chofer'}), 404
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/actualizar-administrativo', methods=['POST'])
def actualizar_administrativo():
    """Actualizar un administrativo existente"""
    try:
        data = request.json
        nombre_original = data.get('nombre_original', '').strip()
        nombre_nuevo = data.get('nombre_nuevo', '').strip().upper()
        
        if not nombre_original or not nombre_nuevo:
            return jsonify({'success': False, 'message': 'Ambos nombres son obligatorios'}), 400
        
        resultado = maestras_manager.actualizar_administrativo_por_nombre(nombre_original, nombre_nuevo)
        
        if resultado:
            return jsonify({'success': True, 'message': 'Administrativo actualizado exitosamente'})
        else:
            return jsonify({'success': False, 'message': 'No se encontró el administrativo'}), 404
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/obtener-chofer')
def obtener_chofer():
    """Obtener datos de un chofer por nombre"""
    try:
        nombre = request.args.get('nombre', '').strip()
        
        if not nombre:
            return jsonify({'success': False, 'message': 'Nombre es obligatorio'}), 400
        
        chofer = maestras_manager.obtener_chofer_por_nombre(nombre)
        
        if chofer:
            return jsonify({'success': True, 'chofer': chofer})
        else:
            return jsonify({'success': False, 'message': 'No se encontró el chofer'}), 404
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

# ========== RUTAS DE RENDICIONES ==========

@app.route('/rendiciones')
@login_required
def rendiciones():
    """Página de gestión de rendiciones de activos"""
    return render_template('rendiciones.html')

@app.route('/api/cargar-rendiciones', methods=['POST'])
@login_required
def cargar_rendiciones():
    """Cargar rendiciones desde archivo Excel"""
    try:
        if 'archivo' not in request.files:
            return jsonify({'success': False, 'message': 'No se recibió ningún archivo'}), 400
        
        archivo = request.files['archivo']
        
        if archivo.filename == '':
            return jsonify({'success': False, 'message': 'No se seleccionó ningún archivo'}), 400
        
        # Verificar extensión
        if not archivo.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'message': 'Solo se permiten archivos Excel (.xlsx, .xls)'}), 400
        
        # Guardar temporalmente
        temp_path = os.path.join(os.path.dirname(__file__), 'temp_rendiciones.xlsx')
        archivo.save(temp_path)
        
        # Procesar
        resultado = rendiciones_manager.cargar_rendiciones_desde_excel(temp_path)
        
        # Eliminar archivo temporal
        try:
            os.remove(temp_path)
        except:
            pass
        
        if resultado['success']:
            mensaje = f"✓ {resultado['registros_cargados']} registros cargados"
            if resultado['duplicados_omitidos'] > 0:
                mensaje += f"\n⚠ {resultado['duplicados_omitidos']} duplicados omitidos"
            if resultado['errores']:
                mensaje += f"\n⚠ {len(resultado['errores'])} errores"
            
            return jsonify({
                'success': True,
                'message': mensaje,
                'registros_cargados': resultado['registros_cargados'],
                'duplicados_omitidos': resultado['duplicados_omitidos'],
                'errores': resultado['errores']
            })
        else:
            return jsonify({'success': False, 'message': resultado.get('error', 'Error desconocido')}), 500
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

@app.route('/api/obtener-rendiciones')
@login_required
def obtener_rendiciones_api():
    """Obtener lista de rendiciones (filtradas)"""
    try:
        filtro = request.args.get('filtro', 'activas')  # 'activas' o 'todas'
        rendiciones = rendiciones_manager.obtener_rendiciones(filtro)
        return jsonify({'success': True, 'rendiciones': rendiciones})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/actualizar-rendicion', methods=['POST'])
@login_required
def actualizar_rendicion_api():
    """Actualizar estado de rendición"""
    try:
        data = request.json
        nro_viaje = data.get('nro_viaje')
        nuevo_estado = data.get('estado')
        
        if not nro_viaje or not nuevo_estado:
            return jsonify({'success': False, 'message': 'Faltan datos requeridos'}), 400
        
        resultado = rendiciones_manager.actualizar_estado_rendicion(nro_viaje, nuevo_estado)
        
        if resultado['success']:
            return jsonify(resultado)
        else:
            return jsonify(resultado), 400
            
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

# ========== SERVIDOR ==========

if __name__ == '__main__':
    # Determinar el directorio base correcto
    if getattr(sys, 'frozen', False):
        # Corriendo como ejecutable empaquetado
        base_dir = os.path.dirname(sys.executable)
    else:
        # Corriendo como script normal
        base_dir = os.path.dirname(os.path.abspath(__file__))
    
    # Crear carpeta pdfs junto al ejecutable o script
    pdfs_dir = os.path.join(base_dir, 'pdfs')
    os.makedirs(pdfs_dir, exist_ok=True)
    
    import socket
    hostname = socket.gethostname()
    local_ip = socket.gethostbyname(hostname)
    
    print("\n" + "="*60)
    print("Sistema de Viajes DHL - Servidor Produccion")
    print("="*60)
    print(f"\nAcceso LOCAL (esta PC):")
    print(f"   http://localhost:5000")
    print(f"   http://127.0.0.1:5000")
    print(f"\nAcceso RED LOCAL (otras PCs en la red):")
    print(f"   http://{local_ip}:5000")
    print(f"\nUsuarios simultaneos: 10")
    print(f"Base de datos: SQLite con WAL mode (viajes.db)")
    print(f"Servidor: Waitress (produccion)")
    print("\n" + "="*60 + "\n")
    
    # Usar Waitress para producción (más estable que Flask development server)
    try:
        from waitress import serve
        serve(app, host='0.0.0.0', port=5000, threads=10)
    except ImportError:
        print("Waitress no instalado. Usando Flask development server...")
        print("   Para producción ejecuta: pip install waitress\n")
        app.run(
            debug=False,
            host='0.0.0.0',
            port=5000,
            threaded=True,
            use_reloader=False
        )
