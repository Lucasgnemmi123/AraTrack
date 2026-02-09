"""
Script para cargar datos desde datos.xlsx a las tablas maestras
"""
import sqlite3
import openpyxl
from datetime import datetime

def cargar_datos_excel():
    """Cargar datos desde el archivo Excel a las tablas maestras"""
    
    print("📁 Abriendo archivo datos.xlsx...")
    try:
        workbook = openpyxl.load_workbook('datos.xlsx')
    except Exception as e:
        print(f"❌ Error al abrir el archivo: {e}")
        return
    
    conn = sqlite3.connect('viajes.db')
    cursor = conn.cursor()
    
    # Reiniciar tablas maestras
    print("\n🔄 Reiniciando tablas maestras...")
    cursor.execute("DROP TABLE IF EXISTS maestras_casinos")
    cursor.execute("DROP TABLE IF EXISTS maestras_choferes")
    cursor.execute("DROP TABLE IF EXISTS maestras_administrativos")
    
    # Recrear tablas
    cursor.execute("""
        CREATE TABLE maestras_casinos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo_costo INTEGER UNIQUE NOT NULL,
            casino TEXT NOT NULL,
            ruta TEXT,
            fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    
    cursor.execute("""
        CREATE TABLE maestras_choferes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT UNIQUE NOT NULL,
            rut TEXT UNIQUE NOT NULL,
            celular TEXT,
            fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    
    cursor.execute("""
        CREATE TABLE maestras_administrativos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT UNIQUE NOT NULL,
            fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    
    print("✅ Tablas maestras reiniciadas")
    
    # Cargar datos de cada pestaña
    
    # 1. Cargar maestras_casinos
    if 'maestras_casinos' in workbook.sheetnames:
        print("\n📊 Cargando maestras_casinos...")
        sheet = workbook['maestras_casinos']
        rows = list(sheet.iter_rows(min_row=2, values_only=True))  # Saltar encabezado
        
        for row in rows:
            if row[0] is not None:  # Si hay codigo_costo
                try:
                    # Convertir a mayúsculas
                    casino = str(row[1]).strip().upper() if row[1] else ''
                    ruta = str(row[2]).strip().upper() if len(row) > 2 and row[2] else ''
                    cursor.execute("""
                        INSERT INTO maestras_casinos (codigo_costo, casino, ruta)
                        VALUES (?, ?, ?)
                    """, (row[0], casino, ruta))
                except sqlite3.IntegrityError as e:
                    print(f"  ⚠️  Código de costo duplicado: {row[0]}")
        
        print(f"  ✅ {len(rows)} registros procesados")
    
    # 2. Cargar maestras_choferes
    if 'maestras_choferes' in workbook.sheetnames:
        print("\n📊 Cargando maestras_choferes...")
        sheet = workbook['maestras_choferes']
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        
        for row in rows:
            if row[0] is not None:  # Si hay nombre
                try:
                    # Convertir a mayúsculas
                    nombre = str(row[0]).strip().upper() if row[0] else ''
                    rut = str(row[1]).strip().upper() if row[1] else ''
                    celular = str(row[2]).strip() if len(row) > 2 and row[2] else ''  # Celular no se convierte a mayúsculas
                    cursor.execute("""
                        INSERT INTO maestras_choferes (nombre, rut, celular)
                        VALUES (?, ?, ?)
                    """, (nombre, rut, celular))
                except sqlite3.IntegrityError as e:
                    print(f"  ⚠️  Nombre o RUT duplicado: {row[0]}")
        
        print(f"  ✅ {len(rows)} registros procesados")
    
    # 3. Cargar maestras_administrativos
    if 'maestras_administrativos' in workbook.sheetnames:
        print("\n📊 Cargando maestras_administrativos...")
        sheet = workbook['maestras_administrativos']
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        
        for row in rows:
            if row[0] is not None:  # Si hay nombre
                try:
                    # Convertir a mayúsculas
                    nombre = str(row[0]).strip().upper() if row[0] else ''
                    cursor.execute("""
                        INSERT INTO maestras_administrativos (nombre)
                        VALUES (?)
                    """, (nombre,))
                except sqlite3.IntegrityError as e:
                    print(f"  ⚠️  Nombre duplicado: {row[0]}")
        
        print(f"  ✅ {len(rows)} registros procesados")
    
    conn.commit()
    
    # Mostrar resumen
    print("\n" + "="*50)
    print("📈 RESUMEN DE CARGA")
    print("="*50)
    
    cursor.execute("SELECT COUNT(*) FROM maestras_casinos")
    print(f"🏢 Casinos/Centros de Costo: {cursor.fetchone()[0]} registros")
    
    cursor.execute("SELECT COUNT(*) FROM maestras_choferes")
    print(f"🚗 Choferes: {cursor.fetchone()[0]} registros")
    
    cursor.execute("SELECT COUNT(*) FROM maestras_administrativos")
    print(f"👤 Administrativos: {cursor.fetchone()[0]} registros")
    
    print("="*50)
    print("✅ Carga completada exitosamente")
    
    conn.close()
    workbook.close()

if __name__ == "__main__":
    cargar_datos_excel()
